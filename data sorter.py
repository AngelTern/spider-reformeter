import pandas as pd
import re
from datetime import datetime
import numpy as np
from collections import Counter
from openpyxl import load_workbook




#ავტვირთოთ მოპოვებული მონაცემედი პანდას დატაფრეიმში
df = pd.read_csv('C:/Users/USER/Desktop/data_extract_sort/scraper/scraper/spiders/123.csv')

#თარიღიდან მოვაშოროთ საათი
df['თარიღი'] = df['თარიღი'].str.replace(r'\d{2}:\d{2}', '', regex=True)

#გადავიყვანოთ თარიღი სტრინგიდან დეიტტაიმ ობიექტში
df['თარიღი'] = pd.to_datetime(df['თარიღი'])

#დავალაგოთ მონაცემები ავთვირთვის თარიღის მიხედვით
df = df.sort_values(by='თარიღი', ascending=True)

#შევქმნათ განახლებული ინდექსაციისთვის ნამპაი ობიექტი
index = np.arange(1, df.shape[0] + 1, 1)

#შევიტანოთ განახლებული ინდექსი დატაფრეიმში
df.index = index

df_2 = pd.read_excel('C:/Users/USER/Desktop/data_extract_sort/insolvency.xlsx', sheet_name='მონაცემები')

from_insolvency = list(df_2['ბმული - e-court-ზე'])

from_scrape = list(df['url'])

from_insolvency = [str(x) for x in from_insolvency]

both = from_scrape + from_insolvency

def get_unique(li):
    c = Counter(li)
    
    result =[key for key,value in c.items() if value == 1]
    
    return result

unique = get_unique(both)

df_unique = pd.DataFrame(unique)

df_unique.to_excel('C:/Users/USER/Desktop/data_extract_sort/unique_urls.xlsx', index = True)
df_to_add = df[df['url'].isin(unique)]

new_column_names = {"მოსამართლე" : "მოსამართლე", "url" : "ბმული - e-court-ზე"}

df_to_add = df_to_add.rename(columns=new_column_names)

#ახალი სვეტების დამატები რომ ექსელს ემთხვეოდეს
#df_to_add[''] = ''
#df_to_add['კომპანიის no'] = ''
#df_to_add[''] = ''
#df_to_add['ბიზნეს სუბიექტი'] = ''
#df_to_add['საქმის ნომერი'] = ''
#df_to_add['ბმული მაცნეზე'] = ''
#df_to_add['საქმის წარმოების დაწყების თარიღი'] = ''
#df_to_add['საქმის წარმოების დაწყების წელი'] = ''
#df_to_add['მეურვე'] = ''
#df_to_add['რეჟიმის მინიჭების თარიღი'] = ''
#df_to_add['რეჟიმი'] = ''
#df_to_add['საქმის მმართველი'] = ''
#df_to_add['მმართველობაში დარჩა'] = ''
#df_to_add['რეჟიმის ცვლილება'] = ''
#df_to_add['რეჟიმის ცვლილების თარიღი'] = ''
#df_to_add['საქმის მმართველი'] = ''
#df_to_add['რეჟიმის ცვლილება II, თარიღი, მმარელი'] = ''
#df_to_add['კრედიტორთა რაოდენობა'] = ''
#df_to_add['სადავო თანხა'] = ''
#df_to_add['სტატუსი'] = ''
#df_to_add['საქმის წარმოების შეწყვეტის/შეჩერების თარიღი'] = ''
#df_to_add['საქმის წარმოების დასრულების თარიღი'] = ''
#df_to_add['გასაჩივრების თარიღი'] = ''
#df_to_add['კომენტარი'] = ''
#df_to_add['ახალი კომენტარი'] = ''
#df_to_add['განხილულია შემდეგი კანონით'] = ''
#df_to_add[''] = ''
#df_to_add['ხანგრძლივობა'] = ''
#df_to_add['დასრულების წელი'] = ''
#df_to_add['შეწყვეტის ან დასრულების თარიღი'] = ''

#new_column_order = ['Unnamed: 0', 'კომპანიის no', 'Unnamed: 2', 'ბიზნეს სუბიექტი', 'საქმის ნომერი', 'მოსამართლე', 'ბმული - e-court-ზე', 'ბმული მაცნეზე', 'საქმის წარმოების დაწყების თარიღი', 'საქმის წარმოების დაწყების წელი', 'მეურვე', 'რეჟიმის მინიჭების თარიღი', 'რეჟიმი', 'საქმის მმართველი', 'მმართველობაში დარჩა', 'რეჟიმის ცვლილება', 'რეჟიმის ცვლილების თარიღი', 'საქმის მმართველი', 'რეჟიმის ცვლილება II, თარიღი, მმართველი', 'კრედიტორთა რაოდენობა', 'სადავო თანხა', 'სტატუსი', 'საქმის წარმოების შეყვეტის/შეჩერების თარიღი', 'საქმის წარმოების დასრულების თარიღი', 'გასაჩივრების თარიღი', 'კომენტარი', 'ახალი კომენტარი', 'განხილულია შემდეგი კანონით', 'Unnamed: 28', 'ხანგრძლიბობა', 'დასრულების წელი', 'შეყვეტის ან დასრულების თარიღი']

#df_to_add = df_to_add.reindex(columns = new_column_order)

df_2.reset_index(drop=True, inplace=True)
df_to_add.reset_index(drop=True, inplace=True)

#df_final = pd.concat([df_2, df_to_add], axis=0, ignore_index=True)


#df_final.to_excel('C:/Users/USER/Desktop/data_extract_sort/final.xlsx')

#path to excel
excel_file = 'C:/Users/USER/Desktop/data_extract_sort/insolvency.xlsx'
#load the workbook
book = load_workbook(excel_file)
#select desired sheet
sheet_name = 'მონაცემები'
writer = pd.ExcelWriter(excel_file, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

# Get the last row of the sheet
sheet = writer.sheets[sheet_name]
last_row = sheet.max_row

# Append DataFrame after the last row
df.to_excel(writer, sheet_name=sheet_name, startrow=last_row, index=False, header=False)

# Save the changes
writer.save()














#df_insolv = pd.DataFrame(from_insolvency)
#df_insolv.to_excel('inslov_urls.xlsx')